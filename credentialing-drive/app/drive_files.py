import csv
import io
from datetime import date, datetime

from fastapi import HTTPException
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

from app.config import (
    CSV_MIME_TYPE, DOCUMENT_MIME_TYPES, GOOGLE_DOCUMENT_MIME_TYPE,
    GOOGLE_SHEETS_MIME_TYPE, MAX_DOCUMENT_BYTES, MAX_SPREADSHEET_BYTES,
    SPREADSHEET_MIME_TYPES,
)



DRIVE_FILE_FIELDS = "id,name,mimeType,size,trashed"


def get_drive_file_metadata(service, file_id):
    return service.files().get(fileId=file_id, fields=DRIVE_FILE_FIELDS).execute()


def download_drive_bytes(service, metadata, export_mime_type=None):
    if export_mime_type:
        request = service.files().export_media(
            fileId=metadata["id"], mimeType=export_mime_type
        )
    else:
        request = service.files().get_media(fileId=metadata["id"])

    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def download_drive_file(
    service,
    file_id,
    allowed_mime_types,
    max_bytes,
    file_kind,
    export_mime_type=None,
    metadata=None,
):
    metadata = metadata or get_drive_file_metadata(service, file_id)
    if metadata.get("trashed"):
        raise HTTPException(status_code=404, detail="Drive file is trashed")
    if metadata.get("mimeType") not in allowed_mime_types:
        raise HTTPException(status_code=415, detail=f"Unsupported {file_kind} format")

    max_megabytes = max_bytes // (1024 * 1024)
    if int(metadata.get("size", 0)) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"{file_kind.capitalize()} exceeds the {max_megabytes} MB processing limit",
        )

    file_bytes = download_drive_bytes(service, metadata, export_mime_type)
    if len(file_bytes) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"{file_kind.capitalize()} exceeds the {max_megabytes} MB processing limit",
        )
    return metadata, file_bytes


def download_drive_document(service, file_id):
    metadata = get_drive_file_metadata(service, file_id)
    export_mime_type = "text/plain" if metadata.get("mimeType") == GOOGLE_DOCUMENT_MIME_TYPE else None
    return download_drive_file(
        service,
        file_id,
        DOCUMENT_MIME_TYPES,
        MAX_DOCUMENT_BYTES,
        "document",
        export_mime_type,
        metadata,
    )


def download_drive_spreadsheet(service, file_id):
    metadata = get_drive_file_metadata(service, file_id)
    export_mime_type = CSV_MIME_TYPE if metadata.get("mimeType") == GOOGLE_SHEETS_MIME_TYPE else None
    return download_drive_file(
        service,
        file_id,
        SPREADSHEET_MIME_TYPES,
        MAX_SPREADSHEET_BYTES,
        "spreadsheet",
        export_mime_type,
        metadata,
    )


def normalize_cell(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value.strip() if isinstance(value, str) else value


def parse_rows(headers, rows):
    raw_headers = []
    used_headers = set()
    for index, header in enumerate(headers, start=1):
        base_header = str(header or "").strip() or f"Column {index}"
        unique_header = base_header
        while unique_header in used_headers:
            unique_header = f"{base_header} ({index})"
        raw_headers.append(unique_header)
        used_headers.add(unique_header)

    parsed_rows = []
    for row_number, values in rows:
        fields = {
            header or f"Column {index}": normalize_cell(value)
            for index, (header, value) in enumerate(zip(raw_headers, values), start=1)
            if value not in (None, "")
        }
        if fields:
            parsed_rows.append({"row_number": row_number, "fields": fields})
    return parsed_rows


def parse_spreadsheet(spreadsheet_bytes, mime_type):
    if mime_type in {CSV_MIME_TYPE, GOOGLE_SHEETS_MIME_TYPE}:
        csv_rows = list(csv.reader(io.StringIO(spreadsheet_bytes.decode("utf-8-sig"))))
        if not csv_rows:
            return []
        return parse_rows(csv_rows[0], enumerate(csv_rows[1:], start=2))

    workbook = load_workbook(io.BytesIO(spreadsheet_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None:
        return []
    return parse_rows(headers, enumerate(rows, start=2))


def split_list(value):
    if not value:
        return []
    return [item.strip() for item in re.split(r"[;,|]", str(value)) if item.strip()]
