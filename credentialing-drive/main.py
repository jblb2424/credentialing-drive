import io
import json
import logging
import os
import csv
import hashlib
import re
import secrets
import uuid
import zipfile
from datetime import date, datetime, timezone
from html import unescape
from xml.etree import ElementTree

from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import RedirectResponse
from google.api_core.exceptions import AlreadyExists
from google.api_core.client_options import ClientOptions
from google import genai
from google.cloud import bigquery, documentai, firestore
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook

app = FastAPI()
logger = logging.getLogger(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
]

CONNECTION_ID = "default"
CONNECTION_COLLECTION = "drive_connections"
EVENT_COLLECTION = "drive_change_events"
PROVIDER_COLLECTION = "providers"
PROVIDER_IDENTITY_COLLECTION = "provider_identities"
BIGQUERY_REPORTING_DATASET = "credentialing_reporting"
BIGQUERY_REPORTING_TABLE = "provider_reporting_events"
PDF_MIME_TYPE = "application/pdf"
JPEG_MIME_TYPE = "image/jpeg"
PNG_MIME_TYPE = "image/png"
GIF_MIME_TYPE = "image/gif"
TIFF_MIME_TYPE = "image/tiff"
BMP_MIME_TYPE = "image/bmp"
WEBP_MIME_TYPE = "image/webp"
DOCX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
GOOGLE_DOCUMENT_MIME_TYPE = "application/vnd.google-apps.document"
TEXT_MIME_TYPES = {"text/plain", "text/rtf", "text/html", "text/markdown"}
CSV_MIME_TYPE = "text/csv"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GOOGLE_SHEETS_MIME_TYPE = "application/vnd.google-apps.spreadsheet"
SPREADSHEET_MIME_TYPES = {CSV_MIME_TYPE, XLSX_MIME_TYPE, GOOGLE_SHEETS_MIME_TYPE}
OCR_DOCUMENT_MIME_TYPES = {
    PDF_MIME_TYPE,
    JPEG_MIME_TYPE,
    PNG_MIME_TYPE,
    GIF_MIME_TYPE,
    TIFF_MIME_TYPE,
    BMP_MIME_TYPE,
    WEBP_MIME_TYPE,
}
TEXT_DOCUMENT_MIME_TYPES = TEXT_MIME_TYPES | {DOCX_MIME_TYPE, GOOGLE_DOCUMENT_MIME_TYPE}
DOCUMENT_MIME_TYPES = OCR_DOCUMENT_MIME_TYPES | TEXT_DOCUMENT_MIME_TYPES
MAX_DOCUMENT_BYTES = 20 * 1024 * 1024
MAX_SPREADSHEET_BYTES = 10 * 1024 * 1024
MAX_GEMINI_INPUT_CHARS = 100_000
DOCUMENT_CATEGORIES = (
    "state_license",
    "board_certificate_or_eligibility_letter",
    "education_training_certificates",
    "dea_registration",
    "ecfmg_certificate",
    "controlled_substance_certificate",
    "cv_or_resume",
    "malpractice_certificate",
    "malpractice_claim_information",
    "hospital_privileges_letter",
    "drivers_license",
    "social_security_card",
    "collaborating_or_supervising_physician_agreement",
    "peer_references",
    "w_9",
    "irs_letter",
    "articles_of_organization",
    "bank_letter_or_voided_check",
    "other",
)

def get_connection_ref():
    database = os.environ.get("FIRESTORE_DATABASE", "healthcare-credentialing")
    client = firestore.Client(database=database)
    return client.collection(CONNECTION_COLLECTION).document(CONNECTION_ID)


def get_connection():
    snapshot = get_connection_ref().get()
    if not snapshot.exists:
        raise HTTPException(status_code=401, detail="Google Drive not connected")
    return snapshot.to_dict()


def update_connection(data):
    get_connection_ref().set(data, merge=True)


def create_flow():
    return Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID"],
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
            }
        },
        scopes=SCOPES,
        redirect_uri=os.environ["GOOGLE_REDIRECT_URI"],
    )


def credentials_to_dict(credentials):
    return {
        "token": credentials.token,
        "refresh_token": credentials.refresh_token,
        "token_uri": credentials.token_uri,
        "client_id": credentials.client_id,
        "client_secret": credentials.client_secret,
        "scopes": list(credentials.scopes),
    }


def get_drive_service(connection=None):
    connection = connection or get_connection()
    credentials_data = connection.get("credentials")
    if not credentials_data:
        raise HTTPException(status_code=401, detail="Google Drive not connected")

    credentials = Credentials(
        token=credentials_data["token"],
        refresh_token=credentials_data["refresh_token"],
        token_uri=credentials_data["token_uri"],
        client_id=credentials_data["client_id"],
        client_secret=credentials_data["client_secret"],
        scopes=credentials_data["scopes"],
    )
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def get_webhook_url():
    url = os.environ.get("GOOGLE_DRIVE_WEBHOOK_URL")
    if not url:
        raise HTTPException(
            status_code=500,
            detail="GOOGLE_DRIVE_WEBHOOK_URL is not configured",
        )
    return url


def get_project_id():
    project_id = os.environ.get("GCP_PROJECT_ID")
    if not project_id:
        raise HTTPException(status_code=500, detail="GCP_PROJECT_ID is not configured")
    return project_id


def download_drive_document(service, file_id):
    metadata = (
        service.files()
        .get(fileId=file_id, fields="id,name,mimeType,size,trashed")
        .execute()
    )
    if metadata.get("trashed"):
        raise HTTPException(status_code=404, detail="Drive file is trashed")
    if metadata.get("mimeType") not in DOCUMENT_MIME_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported document format")

    file_size = int(metadata.get("size", 0))
    if file_size > MAX_DOCUMENT_BYTES:
        raise HTTPException(status_code=413, detail="Document exceeds the 20 MB processing limit")

    if metadata["mimeType"] == GOOGLE_DOCUMENT_MIME_TYPE:
        request = service.files().export_media(fileId=file_id, mimeType="text/plain")
    else:
        request = service.files().get_media(fileId=file_id)

    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    document_bytes = buffer.getvalue()
    if len(document_bytes) > MAX_DOCUMENT_BYTES:
        raise HTTPException(status_code=413, detail="Document exceeds the 20 MB processing limit")
    return metadata, document_bytes


def download_drive_spreadsheet(service, file_id):
    metadata = (
        service.files()
        .get(fileId=file_id, fields="id,name,mimeType,size,trashed")
        .execute()
    )
    if metadata.get("trashed"):
        raise HTTPException(status_code=404, detail="Drive file is trashed")
    if metadata.get("mimeType") not in SPREADSHEET_MIME_TYPES:
        raise HTTPException(status_code=415, detail="Unsupported spreadsheet format")
    if int(metadata.get("size", 0)) > MAX_SPREADSHEET_BYTES:
        raise HTTPException(
            status_code=413, detail="Spreadsheet exceeds the 10 MB processing limit"
        )

    if metadata["mimeType"] == GOOGLE_SHEETS_MIME_TYPE:
        request = service.files().export_media(fileId=file_id, mimeType=CSV_MIME_TYPE)
    else:
        request = service.files().get_media(fileId=file_id)

    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    spreadsheet_bytes = buffer.getvalue()
    if len(spreadsheet_bytes) > MAX_SPREADSHEET_BYTES:
        raise HTTPException(
            status_code=413, detail="Spreadsheet exceeds the 10 MB processing limit"
        )
    return metadata, spreadsheet_bytes


def normalize_cell(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value.strip() if isinstance(value, str) else value


def parse_rows(headers, rows):
    raw_headers = []
    used_headers = set()
    for index, header in enumerate(headers, start=1):
        base_header = str(header or "").strip() or f"Column {index}"
        unique_header = base_header
        while unique_header in used_headers:
            unique_header = f"{base_header} ({index})"
        raw_headers.append(unique_header)
        used_headers.add(unique_header)

    parsed_rows = []
    for row_number, values in rows:
        fields = {
            header or f"Column {index}": normalize_cell(value)
            for index, (header, value) in enumerate(zip(raw_headers, values), start=1)
            if value not in (None, "")
        }
        if fields:
            parsed_rows.append({"row_number": row_number, "fields": fields})
    return parsed_rows


def parse_spreadsheet(spreadsheet_bytes, mime_type):
    if mime_type in {CSV_MIME_TYPE, GOOGLE_SHEETS_MIME_TYPE}:
        csv_rows = list(csv.reader(io.StringIO(spreadsheet_bytes.decode("utf-8-sig"))))
        if not csv_rows:
            return []
        return parse_rows(csv_rows[0], enumerate(csv_rows[1:], start=2))

    workbook = load_workbook(io.BytesIO(spreadsheet_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if headers is None:
        return []
    return parse_rows(headers, enumerate(rows, start=2))


def split_list(value):
    if not value:
        return []
    return [item.strip() for item in re.split(r"[;,|]", str(value)) if item.strip()]


def save_spreadsheet_providers(metadata, providers, document_category):
    for extracted_provider in providers:
        extracted_provider.pop("source_row_numbers", None)
        normalized_provider = normalize_provider_data(extracted_provider)
        upsert_normalized_provider(normalized_provider, metadata, document_category)


def process_drive_spreadsheet_in_memory(service, file_id):
    metadata, spreadsheet_bytes = download_drive_spreadsheet(service, file_id)
    try:
        rows = parse_spreadsheet(spreadsheet_bytes, metadata["mimeType"])
        document_category = classify_document_category(
            metadata, json.dumps(rows, default=str)
        )
        providers = interpret_spreadsheet_with_gemini(rows)
        save_spreadsheet_providers(metadata, providers, document_category)
    finally:
        # Raw spreadsheet bytes are never persisted.
        spreadsheet_bytes = b""
    return {
        "metadata": metadata,
        "document_category": document_category,
        "source_row_count": len(rows),
        "provider_count": len(providers),
    }


def extract_text_with_document_ai(document_bytes, mime_type):
    project_id = get_project_id()
    location = os.environ.get("DOCUMENT_AI_LOCATION", "us")
    processor_id = os.environ.get("DOCUMENT_AI_PROCESSOR_ID")
    if not processor_id:
        raise HTTPException(status_code=500, detail="DOCUMENT_AI_PROCESSOR_ID is not configured")

    client = documentai.DocumentProcessorServiceClient(
        client_options=ClientOptions(api_endpoint=f"{location}-documentai.googleapis.com")
    )
    request = documentai.ProcessRequest(
        name=client.processor_path(project_id, location, processor_id),
        raw_document=documentai.RawDocument(content=document_bytes, mime_type=mime_type),
    )
    document = client.process_document(request=request).document
    return document.text, len(document.pages)


def extract_text_from_docx(document_bytes):
    try:
        with zipfile.ZipFile(io.BytesIO(document_bytes)) as archive:
            document_xml = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile, ElementTree.ParseError):
        raise HTTPException(status_code=422, detail="Invalid DOCX document")

    root = ElementTree.fromstring(document_xml)
    paragraphs = []
    for paragraph in root.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
        text = "".join(
            node.text or ""
            for node in paragraph.findall(
                ".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"
            )
        ).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def extract_text_from_text_document(document_bytes, mime_type):
    if mime_type == DOCX_MIME_TYPE:
        return extract_text_from_docx(document_bytes)

    text = document_bytes.decode("utf-8", errors="replace")
    if mime_type == "text/html":
        text = unescape(re.sub(r"<[^>]+>", " ", text))
    elif mime_type == "text/rtf":
        text = text.replace("\\par", "\n")
        text = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", text)
        text = text.replace("{", " ").replace("}", " ")
    return re.sub(r"[ \t]+", " ", text).strip()


def extract_document_text(document_bytes, mime_type):
    if mime_type in OCR_DOCUMENT_MIME_TYPES:
        return extract_text_with_document_ai(document_bytes, mime_type)
    return extract_text_from_text_document(document_bytes, mime_type), 1


def interpret_text_with_gemini(ocr_text):
    if not ocr_text:
        return {"document_type": "unknown", "summary": "No text extracted"}

    prompt = """You interpret healthcare credentialing documents. Return JSON only with these keys:
document_type, entity_name, group_name, provider, locations, payers, licenses, expiration_dates, and summary.
`provider` must be an object with name, npi, and credentials. `locations`, `payers`, and `licenses` must be arrays.
Use null or empty arrays when a value is not present. Do not infer values that are not supported by the text.

OCR text:
"""
    client = genai.Client(
        vertexai=True,
        project=get_project_id(),
        location=os.environ.get("VERTEX_AI_LOCATION", "global"),
    )
    response = client.models.generate_content(
        model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
        contents=prompt + ocr_text[:MAX_GEMINI_INPUT_CHARS],
        config={"response_mime_type": "application/json", "temperature": 0},
    )
    return response.text


def classify_document_category(metadata, content):
    """Classify each uploaded file once so all resulting revisions share its category."""
    prompt = f"""Classify this healthcare credentialing upload into exactly one category.
Return JSON only in this shape: {{"document_category": "..."}}.
Use only one of these values:
{", ".join(DOCUMENT_CATEGORIES)}.
Choose other when the file does not clearly fit one category. Do not infer a category from
missing information. The file name is {metadata.get("name")!r}.

Document content:
"""
    try:
        client = genai.Client(
            vertexai=True,
            project=get_project_id(),
            location=os.environ.get("VERTEX_AI_LOCATION", "global"),
        )
        response = client.models.generate_content(
            model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
            contents=prompt + content[:MAX_GEMINI_INPUT_CHARS],
            config={"response_mime_type": "application/json", "temperature": 0},
        )
        category = parse_gemini_extraction(response.text).get("document_category")
        return category if category in DOCUMENT_CATEGORIES else "other"
    except Exception:
        logger.exception(
            "Document category classification failed for Drive file_id=%s", metadata.get("id")
        )
        return "other"


def interpret_spreadsheet_with_gemini(rows):
    if not rows:
        return []

    prompt = """You interpret healthcare credentialing onboarding spreadsheets with
unknown column names and layouts. Return JSON only in this shape:
{"providers": [{"source_row_numbers": [2], "document_type": "...",
"entity_name": null, "group_name": null,
"provider": {"name": null, "npi": null, "credentials": null},
"locations": [], "payers": [], "licenses": [], "expiration_dates": [],
"summary": null}]}

Normalize every provider represented in the spreadsheet into this canonical shape.
Use the original row numbers that support each provider in source_row_numbers. A provider
may use multiple rows when the layout requires it. Do not infer values that are not in the
spreadsheet, do not include a provider without a name or NPI, and use null or [] for unknown
values. Preserve all meaningful payer, location, license, and expiration information.

Spreadsheet rows (each object contains an original row_number and raw, client-supplied columns):
"""
    client = genai.Client(
        vertexai=True,
        project=get_project_id(),
        location=os.environ.get("VERTEX_AI_LOCATION", "global"),
    )
    max_batch_chars = MAX_GEMINI_INPUT_CHARS - len(prompt)
    batches = []
    batch = []
    batch_chars = 2
    for row in rows:
        row_json = json.dumps(row, default=str)
        if len(row_json) > max_batch_chars:
            raise HTTPException(status_code=413, detail="Spreadsheet row exceeds the Gemini input limit")
        if batch and batch_chars + len(row_json) + 1 > max_batch_chars:
            batches.append(batch)
            batch = []
            batch_chars = 2
        batch.append(row)
        batch_chars += len(row_json) + 1
    if batch:
        batches.append(batch)

    providers = []
    for batch in batches:
        response = client.models.generate_content(
            model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
            contents=prompt + json.dumps(batch, default=str),
            config={"response_mime_type": "application/json", "temperature": 0},
        )
        providers.extend(parse_gemini_spreadsheet_extraction(response.text, batch))
    return providers


def parse_gemini_extraction(interpretation):
    if isinstance(interpretation, dict):
        return interpretation
    try:
        parsed = json.loads(interpretation)
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="Gemini returned invalid structured output")
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=502, detail="Gemini returned an unexpected response")
    return parsed


def parse_gemini_spreadsheet_extraction(interpretation, source_rows):
    try:
        parsed = json.loads(interpretation) if isinstance(interpretation, str) else interpretation
    except json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="Gemini returned invalid spreadsheet output")
    if not isinstance(parsed, dict) or not isinstance(parsed.get("providers"), list):
        raise HTTPException(status_code=502, detail="Gemini returned an unexpected spreadsheet output")

    valid_row_numbers = {row["row_number"] for row in source_rows}
    providers = []
    for provider in parsed["providers"]:
        if not isinstance(provider, dict):
            raise HTTPException(status_code=502, detail="Gemini returned an invalid provider")
        row_numbers = provider.get("source_row_numbers")
        if not isinstance(row_numbers, list):
            raise HTTPException(status_code=502, detail="Gemini omitted spreadsheet row provenance")
        try:
            row_numbers = sorted({int(row_number) for row_number in row_numbers})
        except (TypeError, ValueError):
            raise HTTPException(status_code=502, detail="Gemini returned invalid spreadsheet row provenance")
        if not row_numbers or not set(row_numbers).issubset(valid_row_numbers):
            raise HTTPException(status_code=502, detail="Gemini returned unknown spreadsheet row provenance")
        provider_profile = provider.get("provider") or {}
        if not isinstance(provider_profile, dict) or not (
            provider_profile.get("name") or provider_profile.get("npi")
        ):
            raise HTTPException(status_code=502, detail="Gemini returned an unidentified provider")
        provider["source_row_numbers"] = row_numbers
        providers.append(provider)
    return providers


def normalized_key(value):
    return re.sub(r"[^a-z0-9]+", "-", str(value or "").lower()).strip("-")


def normalize_provider_data(extraction):
    provider = extraction.get("provider") or {}
    if not isinstance(provider, dict):
        provider = {"name": str(provider)}

    name = provider.get("name") or extraction.get("provider_name")
    npi = provider.get("npi") or extraction.get("npi")
    return {
        "entity_name": extraction.get("entity_name"),
        "group_name": extraction.get("group_name"),
        "provider": {
            "name": name,
            "npi": str(npi) if npi else None,
            "credentials": provider.get("credentials") or extraction.get("credentials"),
        },
        "locations": extraction.get("locations") or [],
        "payers": extraction.get("payers") or [],
        "licenses": extraction.get("licenses") or [],
        "expiration_dates": extraction.get("expiration_dates") or [],
        "summary": extraction.get("summary"),
    }


def resolve_provider_id(client, provider):
    name_key = normalized_key(provider["provider"].get("name"))
    entity_key = normalized_key(provider.get("entity_name"))
    npi = re.sub(r"\D", "", provider["provider"].get("npi") or "")
    if not npi and not name_key:
        return None

    identity_keys = []
    if npi:
        identity_keys.append(f"npi-{npi}")
    if name_key:
        identity_keys.append(f"name-{entity_key or 'unknown'}-{name_key}")

    for identity_key in identity_keys:
        snapshot = client.collection(PROVIDER_IDENTITY_COLLECTION).document(identity_key).get()
        if snapshot.exists:
            provider_id = snapshot.to_dict()["provider_id"]
            for alias_key in identity_keys:
                client.collection(PROVIDER_IDENTITY_COLLECTION).document(alias_key).set(
                    {"provider_id": provider_id}, merge=True
                )
            return provider_id

    # Names and NPIs are aliases used to match later imports. The provider itself
    # receives an opaque Firestore-generated ID.
    provider_id = client.collection(PROVIDER_COLLECTION).document().id
    for identity_key in identity_keys:
        client.collection(PROVIDER_IDENTITY_COLLECTION).document(identity_key).set(
            {"provider_id": provider_id}, merge=True
        )
    return provider_id


def merge_unique(existing, incoming):
    values = list(existing or [])
    for value in incoming or []:
        if value and value not in values:
            values.append(value)
    return values


def reporting_values(values):
    return [
        json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else str(value)
        for value in values or []
        if value is not None
    ]


def sync_provider_to_bigquery(provider_id, provider):
    project_id = get_project_id()
    dataset = os.environ.get("BIGQUERY_REPORTING_DATASET", BIGQUERY_REPORTING_DATASET)
    table_id = f"{project_id}.{dataset}.{BIGQUERY_REPORTING_TABLE}"
    profile = provider["provider"]
    row = {
        "provider_id": provider_id,
        "entity_name": provider.get("entity_name"),
        "group_name": provider.get("group_name"),
        "provider_name": profile.get("name"),
        "npi": profile.get("npi"),
        "credentials": profile.get("credentials"),
        "locations": reporting_values(provider.get("locations")),
        "payers": reporting_values(provider.get("payers")),
        "licenses": reporting_values(provider.get("licenses")),
        "expiration_dates": reporting_values(provider.get("expiration_dates")),
        "synced_at": datetime.now(timezone.utc).isoformat(),
    }
    errors = bigquery.Client(project=project_id).insert_rows_json(
        table_id, [row], row_ids=[f"{provider_id}-{uuid.uuid4()}"]
    )
    if errors:
        raise RuntimeError(f"BigQuery provider sync failed: {errors}")


def provider_changes(existing, updated):
    changes = {}
    for field_name in ("entity_name", "group_name"):
        previous_value = existing.get(field_name)
        current_value = updated.get(field_name)
        if current_value is None or previous_value == current_value:
            continue
        changes[field_name] = {"current": current_value}
        if previous_value is not None:
            changes[field_name]["previous"] = previous_value

    existing_profile = existing.get("provider") or {}
    updated_profile = updated.get("provider") or {}
    profile_changes = {}
    for field_name in ("name", "npi", "credentials"):
        previous_value = existing_profile.get(field_name)
        current_value = updated_profile.get(field_name)
        if current_value is None or previous_value == current_value:
            continue
        profile_changes[field_name] = {"current": current_value}
        if previous_value is not None:
            profile_changes[field_name]["previous"] = previous_value
    if profile_changes:
        changes["provider"] = profile_changes

    for field_name in ("locations", "payers", "licenses", "expiration_dates"):
        added_values = [
            value for value in updated.get(field_name, []) if value not in existing.get(field_name, [])
        ]
        if added_values:
            changes[field_name] = {"added": added_values}
    return changes


def record_provider_revision(provider_ref, metadata, changes, document_category):
    revision = {
        "drive_file_id": metadata["id"],
        "file_name": metadata.get("name"),
        "document_category": document_category,
        "changes": changes,
        "recorded_at": firestore.SERVER_TIMESTAMP,
    }
    provider_ref.collection("revisions").document(f"drive-{metadata['id']}").set(
        revision,
        merge=True,
    )


def scalar_field_changes(changes):
    for field_name in ("entity_name", "group_name"):
        change = changes.get(field_name)
        if change and "current" in change:
            yield field_name, change

    for field_name, change in (changes.get("provider") or {}).items():
        if "current" in change:
            yield f"provider.{field_name}", change


def scalar_revision_changes(changes):
    for field_path, change in scalar_field_changes(changes):
        if "previous" in change:
            yield field_path, change


def source_metadata(metadata):
    return {
        "file_name": metadata.get("name"),
        "drive_file_id": metadata["id"],
    }


def field_provenance_id(field_path):
    return hashlib.sha256(field_path.encode("utf-8")).hexdigest()


def revision_change_for_field(revision, field_path):
    if field_path.startswith("provider."):
        return (revision.get("changes", {}).get("provider", {}) or {}).get(
            field_path.removeprefix("provider.")
        )
    return (revision.get("changes") or {}).get(field_path)


def previous_field_source(provider_ref, field_path, previous_value):
    provenance_ref = provider_ref.collection("field_provenance").document(
        field_provenance_id(field_path)
    )
    provenance = provenance_ref.get().to_dict() or {}
    if provenance.get("value") == previous_value and provenance.get("source"):
        return provenance["source"]

    for revision_snapshot in provider_ref.collection("revisions").stream():
        revision = revision_snapshot.to_dict() or {}
        change = revision_change_for_field(revision, field_path) or {}
        if change.get("current") == previous_value:
            return {
                "file_name": revision.get("file_name"),
                "drive_file_id": revision.get("drive_file_id"),
            }
    return {"source_status": "unavailable"}


def issue_id(field_path, previous_value, current_value):
    values = json.dumps(
        [field_path, previous_value, current_value], sort_keys=True, default=str
    )
    return hashlib.sha256(values.encode("utf-8")).hexdigest()


def record_discrepancies_and_provenance(provider_ref, metadata, changes):
    current_source = source_metadata(metadata)
    for field_path, change in scalar_revision_changes(changes):
        previous_value = change["previous"]
        current_value = change["current"]
        previous_source = previous_field_source(provider_ref, field_path, previous_value)
        provider_ref.collection("issues").document(
            issue_id(field_path, previous_value, current_value)
        ).set(
            {
                "status": "open",
                "field_path": field_path,
                "previous_value": previous_value,
                "previous_source": previous_source,
                "current_value": current_value,
                "current_source": current_source,
                "detected_at": firestore.SERVER_TIMESTAMP,
            },
            merge=True,
        )

    for field_path, change in scalar_field_changes(changes):
        provider_ref.collection("field_provenance").document(
            field_provenance_id(field_path)
        ).set(
            {
                "field_path": field_path,
                "value": change["current"],
                "source": current_source,
                "updated_at": firestore.SERVER_TIMESTAMP,
            },
            merge=True,
        )


def upsert_normalized_provider(provider, metadata, document_category="other"):
    database = os.environ.get("FIRESTORE_DATABASE", "healthcare-credentialing")
    client = firestore.Client(database=database)
    provider_id = resolve_provider_id(client, provider)
    if not provider_id:
        return None

    provider_ref = client.collection(PROVIDER_COLLECTION).document(provider_id)
    existing = provider_ref.get().to_dict() or {}
    existing_profile = existing.get("provider") or {}
    incoming_profile = provider["provider"]
    merged_profile = {
        key: incoming_profile.get(key) or existing_profile.get(key)
        for key in ("name", "npi", "credentials")
    }
    canonical_provider = {
        # Remove the legacy single-document classification from the provider record.
        "document_type": firestore.DELETE_FIELD,
        "entity_name": provider.get("entity_name") or existing.get("entity_name"),
        "group_name": provider.get("group_name") or existing.get("group_name"),
        "provider": merged_profile,
        "locations": merge_unique(existing.get("locations"), provider.get("locations")),
        "payers": merge_unique(existing.get("payers"), provider.get("payers")),
        "licenses": merge_unique(existing.get("licenses"), provider.get("licenses")),
        "expiration_dates": merge_unique(
            existing.get("expiration_dates"), provider.get("expiration_dates")
        ),
    }
    changes = provider_changes(existing, canonical_provider)
    if not changes:
        return provider_id

    provider_ref.set(canonical_provider, merge=True)
    record_provider_revision(provider_ref, metadata, changes, document_category)
    record_discrepancies_and_provenance(provider_ref, metadata, changes)
    try:
        sync_provider_to_bigquery(provider_id, canonical_provider)
    except Exception:
        # Firestore remains the system of record if reporting is temporarily unavailable.
        logger.exception("BigQuery provider sync failed for provider_id=%s", provider_id)
    return provider_id


def save_document_provider(metadata, extraction, document_category):
    provider = normalize_provider_data(extraction)
    return upsert_normalized_provider(provider, metadata, document_category)


def process_drive_document_in_memory(service, file_id):
    metadata, document_bytes = download_drive_document(service, file_id)
    try:
        extracted_text, page_count = extract_document_text(
            document_bytes, metadata["mimeType"]
        )
        document_category = classify_document_category(metadata, extracted_text)
        interpretation = interpret_text_with_gemini(extracted_text)
    finally:
        # Release the raw source document immediately after managed processing.
        document_bytes = b""

    extraction = parse_gemini_extraction(interpretation)
    save_document_provider(metadata, extraction, document_category)
    return {
        "metadata": metadata,
        "document_category": document_category,
        "page_count": page_count,
        "extracted_text": extracted_text,
        "gemini_interpretation": extraction,
    }


def process_drive_changes(service, connection):
    page_token = connection.get("page_token")
    folder_id = connection.get("folder_id")
    if not page_token or not folder_id:
        raise HTTPException(status_code=409, detail="Drive watch is not configured")

    database = os.environ.get("FIRESTORE_DATABASE", "healthcare-credentialing")
    client = firestore.Client(database=database)
    detected_changes = []

    while page_token:
        result = (
            service.changes()
            .list(
                pageToken=page_token,
                spaces="drive",
                fields=(
                    "nextPageToken,newStartPageToken,"
                    "changes(changeType,fileId,removed,file(id,name,mimeType,parents,trashed))"
                ),
            )
            .execute()
        )

        for change in result.get("changes", []):
            file_data = change.get("file") or {}
            if folder_id not in file_data.get("parents", []) or file_data.get("trashed"):
                continue

            file_id = change.get("fileId")
            if not file_id:
                continue

            event = {
                "file_id": file_id,
                "file_name": file_data.get("name"),
                "mime_type": file_data.get("mimeType"),
                "change_type": change.get("changeType"),
                "status": "detected",
            }
            # Use the Drive file ID as an idempotency key across overlapping watch channels.
            event_ref = client.collection(EVENT_COLLECTION).document(f"drive-{file_id}")
            try:
                event_ref.create(event)
            except AlreadyExists:
                continue

            if event["mime_type"] in SPREADSHEET_MIME_TYPES:
                try:
                    result = process_drive_spreadsheet_in_memory(service, file_id)
                except Exception:
                    logger.exception("Spreadsheet import failed for Drive file_id=%s", file_id)
                    event_ref.set({"status": "failed"}, merge=True)
                    detected_changes.append({**event, "status": "failed"})
                    continue

                event_ref.set(
                    {
                        "status": "imported",
                        "source_row_count": result["source_row_count"],
                        "provider_count": result["provider_count"],
                    },
                    merge=True,
                )
                detected_changes.append({**event, "status": "imported"})
                continue

            if event["mime_type"] not in DOCUMENT_MIME_TYPES:
                event_ref.set({"status": "skipped"}, merge=True)
                detected_changes.append({**event, "status": "skipped"})
                continue

            try:
                result = process_drive_document_in_memory(service, file_id)
            except Exception:
                # Preserve no document contents or model output in logs or Firestore.
                logger.exception("Automatic document processing failed for Drive file_id=%s", file_id)
                event_ref.set({"status": "failed"}, merge=True)
                detected_changes.append({**event, "status": "failed"})
                continue

            event_ref.set(
                {
                    "status": "processed",
                    "page_count": result["page_count"],
                    "extracted_character_count": len(result["extracted_text"]),
                    "gemini_response_character_count": len(
                        str(result["gemini_interpretation"])
                    ),
                },
                merge=True,
            )
            logger.info(
                "Automatically processed Drive document file_id=%s pages=%s extracted_characters=%s",
                file_id,
                result["page_count"],
                len(result["extracted_text"]),
            )
            detected_changes.append({**event, "status": "processed"})

        page_token = result.get("nextPageToken")
        if page_token:
            update_connection({"page_token": page_token})
        elif result.get("newStartPageToken"):
            update_connection({"page_token": result["newStartPageToken"]})

    return detected_changes


@app.get("/")
def home():
    return {"status": "ok", "service": "credentialing-drive-test"}


@app.get("/oauth/google/start")
def google_start():
    flow = create_flow()
    authorization_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    update_connection({"oauth_state": state, "code_verifier": flow.code_verifier})
    return RedirectResponse(authorization_url)


@app.get("/oauth/google/callback")
def google_callback(request: Request):
    connection = get_connection()
    if request.query_params.get("state") != connection.get("oauth_state"):
        raise HTTPException(status_code=400, detail="Invalid OAuth state")

    code_verifier = connection.get("code_verifier")
    if not code_verifier:
        raise HTTPException(status_code=400, detail="OAuth authorization has expired")

    flow = create_flow()
    flow.code_verifier = code_verifier
    # Cloud Run terminates TLS before forwarding requests to the container.
    authorization_response = str(request.url.replace(scheme="https"))
    flow.fetch_token(authorization_response=authorization_response)

    update_connection(
        {
            "credentials": credentials_to_dict(flow.credentials),
            "oauth_state": firestore.DELETE_FIELD,
            "code_verifier": firestore.DELETE_FIELD,
        }
    )
    return {"connected": True, "message": "Google Drive connected successfully"}


@app.get("/drive/find-test-folder")
def find_test_folder():
    service = get_drive_service()
    result = (
        service.files()
        .list(
            q=(
                "name='Credentialing Intake' "
                "and mimeType='application/vnd.google-apps.folder' "
                "and trashed=false"
            ),
            spaces="drive",
            fields="files(id,name,parents)",
        )
        .execute()
    )
    folders = result.get("files", [])
    if not folders:
        raise HTTPException(status_code=404, detail="Credentialing Intake folder not found")

    folder = folders[0]
    update_connection({"folder_id": folder["id"]})
    return {"folder": folder}


@app.post("/drive/process/{file_id}")
def process_drive_pdf(file_id: str):
    """Download a Drive document, OCR and interpret it, then release its in-memory bytes."""
    result = process_drive_document_in_memory(get_drive_service(), file_id)
    metadata = result["metadata"]

    logger.info(
        "Processed Drive document file_id=%s name=%s pages=%s extracted_characters=%s",
        metadata["id"],
        metadata.get("name"),
        result["page_count"],
        len(result["extracted_text"]),
    )
    return {
        "processed": True,
        "file": {"id": metadata["id"], "name": metadata.get("name")},
        "page_count": result["page_count"],
        "extracted_text": result["extracted_text"],
        "gemini_interpretation": result["gemini_interpretation"],
    }


@app.post("/drive/watch")
def start_drive_watch():
    connection = get_connection()
    if not connection.get("folder_id"):
        raise HTTPException(status_code=409, detail="Credentialing Intake folder not selected")

    service = get_drive_service(connection)
    page_token = service.changes().getStartPageToken().execute()["startPageToken"]
    channel_id = str(uuid.uuid4())
    channel_token = secrets.token_urlsafe(32)
    response = (
        service.changes()
        .watch(
            pageToken=page_token,
            body={
                "id": channel_id,
                "type": "web_hook",
                "address": get_webhook_url(),
                "token": channel_token,
            },
        )
        .execute()
    )

    update_connection(
        {
            "page_token": page_token,
            "channel_id": channel_id,
            "channel_token": channel_token,
            "resource_id": response["resourceId"],
            "channel_expiration": response.get("expiration"),
        }
    )
    return {
        "watching": True,
        "channel_id": channel_id,
        "expires_at": response.get("expiration"),
    }


@app.post("/drive/watch/renew")
def renew_drive_watch(request: Request):
    # Secret Manager values created from command output can include a trailing newline.
    expected_token = os.environ.get("DRIVE_WATCH_RENEWAL_TOKEN", "").strip()
    provided_token = request.headers.get("x-renewal-token", "")
    if not expected_token or not secrets.compare_digest(provided_token, expected_token):
        raise HTTPException(status_code=401, detail="Invalid renewal token")
    return start_drive_watch()


@app.post("/webhooks/google-drive", status_code=204)
async def google_drive_webhook(request: Request):
    connection = get_connection()
    headers = request.headers
    if not secrets.compare_digest(
        headers.get("x-goog-channel-id", ""), connection.get("channel_id", "")
    ) or not secrets.compare_digest(
        headers.get("x-goog-channel-token", ""), connection.get("channel_token", "")
    ):
        raise HTTPException(status_code=401, detail="Invalid Google Drive channel")

    if headers.get("x-goog-resource-id") != connection.get("resource_id"):
        raise HTTPException(status_code=401, detail="Invalid Google Drive resource")

    changes = process_drive_changes(get_drive_service(connection), connection)
    return Response(status_code=204, headers={"X-Detected-Changes": str(len(changes))})
